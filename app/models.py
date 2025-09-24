import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, date

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
EXCEL_FILE = os.path.join(BASE_DIR, "registros.xlsx")

HEADERS = [
    "FECHA", "CLIENTE", "CELULAR", "ESPECIALIDAD", "MODALIDAD",
    "CUOTA", "TIPO DE CUOTA", "BANCO", "DESTINO", "N° OPERACIÓN",
    "DNI", "CORREO", "GÉNERO", "ASESOR",
]

FIELDS = [
    "fecha", "cliente", "celular", "especialidad", "modalidad",
    "cuota", "tipo_cuota", "banco", "destino", "num_operacion",
    "dni", "correo", "genero", "asesor",
]

def init_excel():
    """Crea el archivo Excel con encabezados si no existe."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

def agregar_registro(data: dict):
    """Agrega una nueva fila de datos al archivo Excel."""
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    fila = [data.get(f, "") for f in FIELDS]
    ws.append(fila)
    wb.save(EXCEL_FILE)

def verificar_duplicado(dni_nuevo, num_operacion_nuevo):
    """
    Verifica si un DNI o N° de Operación ya existe en el Excel.
    Devuelve un mensaje de error si se encuentra un duplicado, o None si no hay.
    """
    if not os.path.exists(EXCEL_FILE):
        return None
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    dni_col_idx = FIELDS.index("dni")
    op_col_idx = FIELDS.index("num_operacion")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(dni_col_idx, op_col_idx):
            continue
        dni_existente = str(row[dni_col_idx] or "").strip()
        num_op_existente = str(row[op_col_idx] or "").strip()
        if dni_existente == dni_nuevo:
            return f"Error: El DNI '{dni_nuevo}' ya se encuentra registrado."
        if num_operacion_nuevo and num_op_existente == num_operacion_nuevo:
            return f"Error: El N° de Operación '{num_operacion_nuevo}' ya se encuentra registrado."
    return None

def buscar_registros(query):
    """Busca registros por DNI o nombre de cliente."""
    resultados = []
    if not query or not os.path.exists(EXCEL_FILE):
        return resultados
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True), start=2):
        dni = str(row[FIELDS.index("dni")] or "").lower()
        cliente = str(row[FIELDS.index("cliente")] or "").lower()
        if query in dni or query in cliente:
            resultados.append((row, idx))
    return resultados

def obtener_datos_fila(fila_idx):
    """Obtiene los datos de una fila específica para la edición."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    fila_valores = [ws.cell(row=fila_idx, column=col).value for col in range(1, len(FIELDS) + 1)]
    data = {field: _format_cell_value(val) for field, val in zip(FIELDS, fila_valores)}
    return data

def actualizar_fila(fila_idx, form_data):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for col_idx, field in enumerate(FIELDS, start=1):
        valor = form_data.get(field, "")
        ws.cell(row=fila_idx, column=col_idx).value = valor
    wb.save(EXCEL_FILE)

def eliminar_fila(fila_idx):
    """Elimina una fila por su índice."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.delete_rows(fila_idx)
        wb.save(EXCEL_FILE)
        return True
    except Exception:
        return False

def _format_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)

# La función de reporte ahora acepta fechas como parámetros
def generar_reporte_asesores(start_date_str=None, end_date_str=None):
    """
    Procesa el archivo Excel para generar un reporte detallado de ventas 
    por asesor, aplicando un filtro de fechas si se proporciona.
    """
    if not os.path.exists(EXCEL_FILE):
        return {}
    
    # Convertir fechas de string a objetos datetime
    start_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            pass # Ignorar si el formato es incorrecto
    
    end_date = None
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            pass
        
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    reporte = {}
    asesor_col_idx = FIELDS.index("asesor")
    cuota_col_idx = FIELDS.index("cuota")
    fecha_col_idx = FIELDS.index("fecha")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(asesor_col_idx, cuota_col_idx, fecha_col_idx):
            continue

        asesor_nombre = str(row[asesor_col_idx] or "Sin Asesor").strip()
        fecha_valor = row[fecha_col_idx]

        # Validar la fecha para el filtro
        if isinstance(fecha_valor, datetime):
            current_date = fecha_valor
        elif isinstance(fecha_valor, str):
            try:
                current_date = datetime.strptime(fecha_valor, '%Y-%m-%d')
            except ValueError:
                continue
        else:
            continue
        
        # Aplicar el filtro de fechas
        if start_date and current_date < start_date:
            continue
        if end_date and current_date > end_date:
            continue

        try:
            cuota_valor = float(row[cuota_col_idx])
        except (ValueError, TypeError):
            continue

        if asesor_nombre not in reporte:
            reporte[asesor_nombre] = {
                'total_asesor': 0.0,
                'registros_asesor': 0
            }

        reporte[asesor_nombre]['total_asesor'] += cuota_valor
        reporte[asesor_nombre]['registros_asesor'] += 1
            
    return reporte